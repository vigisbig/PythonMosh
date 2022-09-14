import openpyxl as xl
wb = xl.load_workbook('c:\\Users\panka\OneDrive\Desktop\File.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']

for row in range(2,sheet.max_row+1):
    cells_to_be_operated1 = sheet.cell(row,1)
    cells_to_be_operated2 = sheet.cell(row,2)
    value1 = cells_to_be_operated1.value
    value2 = cells_to_be_operated2.value
    cells_to_go = sheet.cell(row,4)
    cells_to_go.value = value1 * value2
wb.save('c:\\Users\panka\OneDrive\Desktop\Filenew.xlsx')
    
    