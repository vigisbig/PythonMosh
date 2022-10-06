import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(filename):                                          #   Putting all code in a function so that it can be called and files can be processed 
    wb = xl.load_workbook(filename)                                      #   using .load_workbook method and specifying file path/name to load, assigning to a variable wb
    sheet = wb['Sheet1']                                                 #   Using square bracket to access sheet of the file, and calling variable as sheet
    #cell = sheet['a1']                                                      #   This format is used to access cell
    #cell = sheet.cell(1,1)                                                  #   Or you can use this method to access cell                                                            
    #print(cell.value)                                                       #   Printing the value of selected cell
    #print(sheet.max_column)                                                 #   Printing the number of columns, it will also work for rows max_rows
    for row in range(2,sheet.max_row+1):
        cell1 = sheet.cell(row,2)                                            #   Creating variable cell1 and specifying that it's from row (from for loop) and second column
        #print(cell.value)                                                   #   Can be used to check if you are on the right track                                   
        cell2 = sheet.cell(row,3)                                            #   Creating variable cell2 and specifying that it's from row (from for loop) and third column
        corrected_price = cell1.value*cell2.value                            #   Creating new variable to store the value that's created by multiplying values of cell1 and cell2
        corrected_price_cell = sheet.cell(row,4)                             #   specifying the cells we want to use for resulting calculation values 
        corrected_price_cell.value = corrected_price                         #   Assigning the values 
    values = Reference(sheet,min_row=1, max_row=sheet.max_row,min_col=2,max_col=4)      # creating an object of class Reference 
    chart = BarChart()                                                                  # creating an object of class BarChart
    chart.add_data(values)                                                              # Using add_data method with Bar_Chart class object to specify values we want to bar chart 
    sheet.add_chart(chart,'f2')                                                         # Using add_chart method to create chart and specifying the location start point as f2
    wb.save(filename)                                                                   # Saving file